#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Version : 1.0
# @Time    : 2017/11/25
# @Author  : Yummy

import gc
from openpyxl import load_workbook


######读取excel表单元格数据，Return 一个二维字典数据#######
def DataRead(Sheetname):
    wb = load_workbook(Sheetname)
    sheetnames = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetnames[0])
    DictSku = {}
    DictLine = {}
    i = 1
    for sku in sheet["B"]:
            for cell in sheet[i]:
                if (cell.value):
                    CellKey = cell.coordinate
                    DictLine[CellKey] = cell.value
                else:
                    continue
            i = i + 1
            print(i)
            DictSku[sku.value] = DictLine
            gc.collect()
    wb.close()
    return DictSku


#####  数据写入  ##########
def DataWriting(SheetWriting,SheetReading):
    wb = load_workbook(SheetWriting)
    sheetnames = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetnames[0])
    Data = DataRead(SheetReading)
    for sku in sheet['B']:
        key = sku.value
        if (key in Data.keys()):
            for key,value in Data[key].items():
                if (sheet[key].value is  None):
                    print (key)
                    print (value)
                    sheet[key] = value
                else:
                    continue
        else:
            continue
    wb.close()
    wb.save(SheetWriting)
print ("Start")
DataWriting("writing3.xlsx","reading3.xlsx")
exit()

